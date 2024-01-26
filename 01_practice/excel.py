from openpyxl import load_workbook
from openpyxl import Workbook

load_wb = load_workbook("D:\\03.Test\\v620\\9R_231207\\ori_ExynosAutoV620_Validation_IR231207_070011.xlsx")
number_list = []  # Initialize an empty list
tc_number = []
category = []
BL = []
BA = []
LA = []
tree_data = load_wb['Add']
number = 0
for i in range(6, 228):
    number_list.append(load_wb['Validation Result'][f'A{i}'].value)  # Append values to the list
    tc_number.append(load_wb['Validation Result'][f'B{i}'].value)  # Append values to the list
    category.append(load_wb['Validation Result'][f'C{i}'].value)  # Append values to the list
    BL.append(load_wb['Validation Result'][f'J{i}'].value)  # Append values to the list
    BA.append(load_wb['Validation Result'][f'K{i}'].value)  # Append values to the list
    LA.append(load_wb['Validation Result'][f'L{i}'].value)  # Append values to the list
    tree_data.cell(row=number+1, column=2).value = number_list[number]
    tree_data.cell(row=number+1, column=3).value = tc_number[number]
    tree_data.cell(row=number+1, column=4).value = category[number]
    tree_data.cell(row=number+1, column=5).value = BL[number]
    tree_data.cell(row=number+1, column=6).value = BA[number]
    tree_data.cell(row=number+1, column=7).value = LA[number]

    print(number_list[number])
    print(tc_number[number])
    print(category[number])
    print(BL[number])
    print(BA[number])
    print(LA[number])

    number += 1

load_wb.save("D:\\03.Test\\v620\\9R_231207\\ori_ExynosAutoV620_Validation_IR231207_070011.xlsx")




